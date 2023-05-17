from xml.dom.minidom import parse
import xml.dom.minidom
import re

from openpyxl import Workbook
workbook = Workbook() #create a workbook
worksheet = workbook.active #active the worksheet

#set the column names
worksheet['A1'] = 'id'
worksheet['B1'] = 'name'
worksheet['C1'] = 'definition'
worksheet['D1'] = 'childnodes'
row_num = 1

DOMTree = xml.dom.minidom.parse("go_obo.xml")
obo = DOMTree.documentElement
GO_term = obo.getElementsByTagName('term')

#define a function to count the childnodes
def count_childnodes(id):
  node_num = 0
  for term_for_childnodes in GO_term:
     new_id = term_for_childnodes.getElementsByTagName('id')[0]
     is_a_childnodes = term_for_childnodes.getElementsByTagName('is_a')
     for is_a_sub in is_a_childnodes:
              childnode = is_a_sub.childNodes[0].data
              if id == childnode:  node_num = node_num + count_childnodes(new_id.childNodes[0].data) + 1
  return node_num

for term in GO_term:
 defstr = term.getElementsByTagName('defstr')[0]
 #check whether there is a word "autophagosome" in the text in <defstr>
 if re.search('autophagosome',str(defstr.childNodes[0].data)):
  #get the id
  id = term.getElementsByTagName('id')[0]
  #get the name
  name = term.getElementsByTagName('name')[0]
  #write down the data into the sheet
  row_num = row_num + 1
  worksheet.cell(row=row_num, column=1).value = id.childNodes[0].data
  worksheet.cell(row=row_num, column=2).value = name.childNodes[0].data
  worksheet.cell(row=row_num, column=3).value = defstr.childNodes[0].data
  #count the number of childnodes 
  node_num = count_childnodes(id.childNodes[0].data)
  worksheet.cell(row=row_num, column=4).value = node_num

#save the workbook
workbook.save("autophagosome.xlsx")
